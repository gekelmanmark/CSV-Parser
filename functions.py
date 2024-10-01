from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from collections import OrderedDict
import datetime
import csv

# A data class that helps retain the information needed from each row in one variable
@dataclass
class CellData:
    device_name: str # The device names
    msg_sent_date: str # Date the msg was sent (datetime object)
    msg_received_date: str # Date the msg was received (datetime object)
    success: str # Success message given from OnPortal (boolean)

@dataclass
class OutPingData:
    device_name: str
    num_true: int
    num_false: int
    response_time: list

# Format of the datetime string from the .csv
ALPHA_TEST_datetime_format = '%m/%d/%Y %H:%M:%S %p'
current_format = ALPHA_TEST_datetime_format


# datetime(year, month, day, hour=0, minute=0, second=0, microsecond=0, tzinfo=None, *, fold=0)
alpha_test_begining_date = datetime.datetime(2024, 7, 22).date() # Start of the alpha testing
alpha_test_middle_date = datetime.datetime(2024, 7, 29).date() # Middle of the alpha testing
alpha_test_end_date = datetime.datetime(2024, 8, 5).date() # End of alpha testing ( 2 weeks )

grand_begining_date = datetime.datetime(2024, 9, 23).date() # Start of the alpha testing
grand_end_date = datetime.datetime(2024, 9, 29).date() # End of alpha testing ( 2 weeks )

# Checks if a worksheet is empty
def WorksheetEmpty(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False
    return True

def CreateChart(ws):
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Bar Chart"
    chart1.y_axis.title = 'Test number'
    chart1.x_axis.title = 'Sample length (mm)'

    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "G1")

def getPings(struct_list, worksheet):
    dict = OrderedDict()
    for struct in struct_list:
        if struct.device_name not in dict: # If the device name has not been added to the dictionary yet
            datetime_msg_sent = datetime.datetime.strptime(struct.msg_sent_date, current_format)
            datetime_msg_received = datetime.datetime.strptime(struct.msg_received_date, current_format)
            if (datetime_msg_received.date() >= grand_begining_date) and (datetime_msg_received.date() <= grand_end_date): # If the msg. received date is within the time period
                dict[struct.device_name] = OutPingData(struct.device_name, 0, 0, [])
                if struct.success == 'TRUE' or struct.success == 'True': # If the ping returned successful
                    dict[struct.device_name].num_true = 1
                    dict[struct.device_name].num_false = 0
                    datetime_time_difference = datetime_msg_received - datetime_msg_sent
                    dict[struct.device_name].response_time.append(datetime_time_difference.total_seconds())
                else: # The Ping was not successful
                    dict[struct.device_name].num_false = 1
                    dict[struct.device_name].num_true = 0
        else: # The device name already exists within the dictionary
            datetime_msg_sent = datetime.datetime.strptime(struct.msg_sent_date, current_format)
            datetime_msg_received = datetime.datetime.strptime(struct.msg_received_date, current_format)
            if (datetime_msg_received.date() >= grand_begining_date) and (datetime_msg_received.date() <= grand_end_date): # If the msg. received date is within the time period
                if struct.success == 'TRUE' or struct.success == 'True': # If the ping returned successful
                    dict[struct.device_name].num_true = dict[struct.device_name].num_true + 1
                    datetime_time_difference = datetime_msg_received - datetime_msg_sent
                    dict[struct.device_name].response_time.append(datetime_time_difference.total_seconds())
                else: # The Ping was not successful
                    dict[struct.device_name].num_false = dict[struct.device_name].num_false + 1
    if len(dict) > 0:
        list_column_names = ['Device', 'Success Pings', 'Fail Pings', 'Avg. Response Time (s)', 'Avg. Success (%)']
        worksheet.append(list_column_names)

    for device, value in dict.items():
        list_row = [float(device)] # Device name
        list_row.append(value.num_true) # Success Pings
        list_row.append(value.num_false) # Fail Pings
        if len(value.response_time) == 0: # Avg. Response Time (s)
            list_row.append(0)
        else:
            total_response_time = 0
            for i in value.response_time:
                total_response_time = total_response_time + i
            list_row.append(round(total_response_time / len(value.response_time), 2))
        list_row.append(round((value.num_true / (value.num_true + value.num_false) * 100), 2)) # Avg. Success (%)
        worksheet.append(list_row)

def read_csv(file_name):
    dict_usable_data = {} # Creating a Dict. object with just the data that we need to extract
    with open(file_name, newline='') as csvfile:
        reader = csv.DictReader(csvfile) # Reading in the data into a dictionary object
        for row in reader: # For each row in the CSV file...
            if row['Info'] not in dict_usable_data:
                dict_usable_data[row['Info']] = [] # This will add a key for every unique Info response in AlphaTest.csv
            # Adding needed data from row into dict_usable_data dictionary 
            if "âŠ™" in str(row['DeviceName']):
                row['DeviceName'] = str(row['DeviceName']).replace("âŠ™", ".")
            if str(row['DeviceName'])[0].isdigit():
                dict_usable_data[row['Info']].append(CellData(row['DeviceName'], 
                                                row['MessageSentDate'], 
                                                row['MessageReceivedDate'],
                                                row['Success']))
    return dict_usable_data

def Mesh_Status(usable_data, main_wb, device_list, sheet_name):
    # Extracting and breaking apart the data further...
    list_mesh_status = usable_data['Mesh Status'] # Extracting 'Mesh Status' (list of CellData objects)
    dict_device_ping_dates = {} # All the dates every device pinged its status

    list_mesh_status.sort(key=lambda x: x.device_name) # Sorting the list by device name
    # mesh_wb = Workbook() # Creating new file (workbook)
    mesh_ws = main_wb.create_sheet(sheet_name)

    current_device = ""
    for struct in list_mesh_status:
        datetime_recieved_date = datetime.datetime.strptime(struct.msg_received_date, current_format).date()
        # If we come accross a new device's data
        if struct.device_name != current_device: 
            current_device = struct.device_name
            dict_device_ping_dates[current_device] = []
            if (datetime_recieved_date >= grand_begining_date) and (datetime_recieved_date <= grand_end_date):
                dict_device_ping_dates[current_device].append(datetime_recieved_date)
        # If the device is not new, but there is a new day the device pinged    
        elif (datetime_recieved_date not in dict_device_ping_dates[current_device]) and (datetime_recieved_date >= grand_begining_date) and (datetime_recieved_date <= grand_end_date): 
            dict_device_ping_dates[current_device].append(datetime_recieved_date)
        else: # Else -> continue to next struct in the list
            continue

    # Calculating number of days between the start and finish of the alpha test
    # total_num_days = alpha_test_end_date - alpha_test_begining_date
    total_num_days = grand_end_date - grand_begining_date

    column_names = ['Device', 'Successful Response(s)', 'Missing Response(s)', 'Avg. Success (%)']
    mesh_ws.append(column_names)
    
    for device in device_list:
        row_list = []
        row_list.append(float(device)) # Device
        if device in dict_device_ping_dates:
            row_list.append(int(len(dict_device_ping_dates[device]))) # Success Pings
            row_list.append(int((total_num_days.days + 1) - len(dict_device_ping_dates[device]))) # Total days (inclusive) minus the amount of pings counted for this device
            row_list.append(round((len(dict_device_ping_dates[device]) / (total_num_days.days + 1)) * 100, 2)) # Avg. Success (%)
        else:
            row_list.append('N/A') # Success Pings
            row_list.append(total_num_days.days + 1) # Missing Pings
            row_list.append(0) # Avg. Success (%)
        mesh_ws.append(row_list)

def OutgoingPingRequestNotification(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingPingRequestNotification'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def export_msg_ctr_sync_response_to_file(usable_data, sheet_name):
    # Extracting and breaking apart the data further...
    list_sync_response = usable_data['Msg Ctr Sync Response'] # Extracting 'Msg Ctr Sync Response' (list of CellData objects)
    for row_data in list_sync_response:
        print(f"{row_data.device_name} ---> {row_data.success}")

def OutgoingMeshDeviceDiagnosticRequest(usable_data, main_wb, sheet_name):
    # Extracting and breaking apart the data further...
    list_diagnostic_request_data = usable_data['OutgoingMeshDeviceDiagnosticRequest'] # Extracting 'OutgoingMeshDeviceDiagnosticRequest' (list of CellData objects)
    list_diagnostic_request_data.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(list_diagnostic_request_data, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def Open(usable_data, main_wb, sheet_name):
    struct_list = usable_data['Open'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def Close(usable_data, main_wb, sheet_name):
    struct_list = usable_data['Close'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OpenAndClose(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OpenAndClose'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanReadAuditRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanReadAuditRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanAuditPointerRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanAuditPointerRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingTraceRtRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingTraceRtRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanStateRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanStateRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanDSTReadRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanDSTReadRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingMeshEventNodeIdsRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingMeshEventNodeIdsRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def LockBlockUnblock(usable_data, main_wb, sheet_name):
    struct_list = usable_data['LockBlockUnblock'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingMasterUserCancelRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingMasterUserCancelRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingCardErrorDiagnoseRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingCardErrorDiagnoseRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingMeshStatusWindowRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingMeshStatusWindowRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanProgrammingRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanProgrammingRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanCalendarRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanCalendarRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanAutomaticChangesRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanAutomaticChangesRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanShiftsRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanShiftsRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanMasterCardKeycodesRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanMasterCardKeycodesRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanGuestCardKeycodeRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanGuestCardKeycodeRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanSpecialCardKeycodesRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanSpecialCardKeycodesRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanRtcRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanRtcRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanPriorityMsgConfigRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanPriorityMsgConfigRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanGeneralRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanGeneralRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockDiagnosticGetRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockDiagnosticGetRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingLockingPlanRTCReadRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingLockingPlanRTCReadRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def OutgoingMFGDateRequest(usable_data, main_wb, sheet_name):
    struct_list = usable_data['OutgoingMFGDateRequest'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def LockingPlanRead(usable_data, main_wb, sheet_name):
    struct_list = usable_data['LockingPlanRead'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def ModifyStayDeviceSource(usable_data, main_wb, sheet_name):
    struct_list = usable_data['ModifyStayDeviceSource'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]

def ModifyStayDeviceDestination(usable_data, main_wb, sheet_name):
    struct_list = usable_data['ModifyStayDeviceDestination'] # Extracting 'Open' (list of CellData objects)
    struct_list.sort(key=lambda x: (x.device_name)) # Sorting the list by device name
    ws = main_wb.create_sheet(sheet_name)
    getPings(struct_list, ws)
    if WorksheetEmpty(ws):
        del main_wb[sheet_name]